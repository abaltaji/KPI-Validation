"""Module for generating reports (Excel, Google Sheets)."""

import json
import os
import traceback
from json import JSONDecodeError
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import gspread
from google.oauth2.service_account import Credentials

# =============================================================================
# SECTION 1: DATA PROCESSING HELPERS (SHARED LOGIC)
# =============================================================================


def _normalize_rows(rows: list[dict]):
    """Ensure rows have PRG_PAR keys required for validation and comparison."""
    for r in rows:
        if "PRG_PAR_Area" not in r: r["PRG_PAR_Area"] = r.get("area")
        if "PRG_PAR_UseRatio" not in r: r["PRG_PAR_UseRatio"] = r.get("use_ratio")
        if "PRG_PAR_ResourceConsRatio" not in r: r["PRG_PAR_ResourceConsRatio"] = r.get("resource_cons_ratio")
        if "PRG_PAR_GeometryWeight" not in r: r["PRG_PAR_GeometryWeight"] = r.get("geometry_weight")
        if "PRG_PAR_MeanDistToExit" not in r: r["PRG_PAR_MeanDistToExit"] = r.get("mean_dist_to_exit")
        if "PRG_PAR_IdealDistToExit" not in r: r["PRG_PAR_IdealDistToExit"] = r.get("ideal_dist_to_exit")


def _prepare_raw_data(rows: list[dict]) -> tuple[list[list], dict]:
    """
    Process rows to generate the raw data table and grand totals.
    Returns (table_rows, grand_totals).
    """
    aggregated_data = defaultdict(lambda: defaultdict(float))
    
    for r in rows:
        program = r.get("program")
        if not program or program == "Unspecified":
            continue

        aggregated_data[program]["count"] += 1
        aggregated_data[program]["area"] += r.get("area", 0.0)
        aggregated_data[program]["use_ratio"] = r.get("use_ratio", 0.0)
        aggregated_data[program]["resource_cons_ratio"] = r.get("resource_cons_ratio", 0.0)
        aggregated_data[program]["geometry_weight"] += r.get("geometry_weight", 0.0)
        aggregated_data[program]["mean_dist_to_exit"] += r.get("mean_dist_to_exit", 0.0)
        aggregated_data[program]["ideal_dist_to_exit"] = r.get("ideal_dist_to_exit", 0.0)

    grand_totals = defaultdict(float)
    table_rows = []
    
    for program, data in sorted(aggregated_data.items()):
        grand_totals["area"] += data["area"]
        grand_totals["geometry_weight"] += data["geometry_weight"]
        
        count = data["count"] if data["count"] > 0 else 1
        avg_mean_dist = data["mean_dist_to_exit"] / count
        
        table_rows.append([
            program,
            data["area"],
            data["use_ratio"],
            data["resource_cons_ratio"],
            data["geometry_weight"],
            avg_mean_dist,
            data["ideal_dist_to_exit"]
        ])
        
    return table_rows, grand_totals


def _prepare_summary_data(rows: list[dict]) -> dict:
    """
    Calculate summary statistics (Area per Tower, Level, Program, Matrix).
    """
    total_area = sum(r["area"] for r in rows)
    area_per_tower = defaultdict(float)
    area_per_level = defaultdict(float)
    area_per_program = defaultdict(float)
    matrix = defaultdict(lambda: defaultdict(float))

    for r in rows:
        tower = r["tower"] or "Unknown"
        level = r["level"] or "Unspecified"
        program = r["program"] or "Unspecified"
        area = r["area"]

        area_per_tower[tower] += area
        area_per_level[level] += area
        area_per_program[program] += area
        matrix[level][tower] += area
        
    return {
        "total_area": total_area,
        "area_per_tower": area_per_tower,
        "area_per_level": area_per_level,
        "area_per_program": area_per_program,
        "matrix": matrix
    }

# =============================================================================
# SECTION 2: LOGIC - COMPARISON & VALIDATION
# =============================================================================

def _compare_versions(current_rows: list[dict], previous_rows: list[dict]) -> list[dict]:
    """
    Compare current and previous versions to identify changes.
    
    Returns a list of comparison records showing:
    - Program name
    - Field that changed
    - Previous value
    - Current value
    - Status (Added, Modified, Removed)
    """
    comparison_data = []
    
    # Create lookups for O(1) access
    current_by_program = {row["program"]: row for row in current_rows if row.get("program") and row["program"] != "Unspecified"}
    previous_by_program = {row["program"]: row for row in previous_rows if row.get("program") and row["program"] != "Unspecified"}
    
    fields_to_compare = [
        "PRG_PAR_Area",
        "PRG_PAR_UseRatio",
        "PRG_PAR_ResourceConsRatio",
        "PRG_PAR_GeometryWeight",
        "PRG_PAR_MeanDistToExit",
        "PRG_PAR_IdealDistToExit",
    ]
    
    # Check for new and modified programs
    for program, current_data in current_by_program.items():
        if program not in previous_by_program:
            # New program added
            comparison_data.append({
                "program": program,
                "field": "Program",
                "previous_value": "N/A",
                "current_value": "New Program",
                "status": "Added",
                "is_changed": True
            })
        else:
            # Check for field modifications
            previous_data = previous_by_program[program]
            for field in fields_to_compare:
                current_val = current_data.get(field, 0.0)
                previous_val = previous_data.get(field, 0.0)
                
                # Convert to float for comparison
                try:
                    current_val_num = float(current_val) if current_val else 0.0
                    previous_val_num = float(previous_val) if previous_val else 0.0
                except (ValueError, TypeError):
                    current_val_num = str(current_val)
                    previous_val_num = str(previous_val)
                
                if current_val_num != previous_val_num:
                    comparison_data.append({
                        "program": program,
                        "field": field,
                        "previous_value": previous_val,
                        "current_value": current_val,
                        "status": "Modified",
                        "is_changed": True
                    })
    
    # Check for removed programs
    for program, previous_data in previous_by_program.items():
        if program not in current_by_program:
            comparison_data.append({
                "program": program,
                "field": "Program",
                "previous_value": "Removed Program",
                "current_value": "N/A",
                "status": "Removed",
                "is_changed": True
            })
    
    return sorted(comparison_data, key=lambda x: (x["program"], x["field"]))


def _create_comparison_sheet(workbook: Workbook, current_rows: list[dict], previous_rows: list[dict] = None):
    """
    Create a 'Version Comparison' sheet showing changes between versions.
    """
    if previous_rows is None or len(previous_rows) == 0:
        return  # Skip if no previous version to compare
    
    comparison_data = _compare_versions(current_rows, previous_rows)
    
    if not comparison_data:
        return  # No changes detected
    
    comparison_sheet = workbook.create_sheet("Version Comparison")
    
    # Create Data Validation for Status column
    dv = DataValidation(type="list", formula1='"Added,Modified,Removed"', allow_blank=True)
    comparison_sheet.add_data_validation(dv)

    # Define borders
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    light_blue_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
    
    # Headers
    headers = ["Program", "Field", "Previous Value", "Current Value", "Status"]
    comparison_sheet.append(headers)
    
    # Style headers - match Program_KPI Parameters color
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in comparison_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
    
    # Color fills for different statuses
    added_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    modified_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
    removed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    
    status_colors = {
        "Added": added_fill,
        "Modified": modified_fill,
        "Removed": removed_fill
    }
    
    # Add data rows with gaps between program changes
    previous_program = None
    for comp in comparison_data:
        # Add a light blue gap row when program name changes
        if previous_program is not None and comp["program"] != previous_program:
            # Add empty row (white separator)
            comparison_sheet.append([None, None, None, None, None])
            # No borders or fill for the gap row to make it a clean separator
        
        row = [
            comp["program"],
            comp["field"],
            comp["previous_value"],
            comp["current_value"],
            comp["status"]
        ]
        comparison_sheet.append(row)
        
        # Apply highlighting
        row_idx = comparison_sheet.max_row
        status_fill = status_colors.get(comp["status"], PatternFill())
        status_cell = comparison_sheet.cell(row=row_idx, column=5)
        
        # Add Status cell to validation
        dv.add(status_cell)
        
        # Apply static highlighting to data columns (A-D)
        for col_idx in range(1, 5):
            cell = comparison_sheet.cell(row=row_idx, column=col_idx)
            if comp["is_changed"]:
                cell.fill = status_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
        
        # Apply formatting to Status column (E) - Fill handled by Conditional Formatting below
        status_cell.alignment = Alignment(horizontal="left", vertical="center")
        status_cell.border = thin_border
        
        previous_program = comp["program"]
    
    # Apply Conditional Formatting to Status Column (E)
    # This ensures the color changes if the user changes the dropdown value
    range_string = f"E2:E{comparison_sheet.max_row}"
    comparison_sheet.conditional_formatting.add(range_string, CellIsRule(operator='equal', formula=['"Added"'], stopIfTrue=True, fill=added_fill))
    comparison_sheet.conditional_formatting.add(range_string, CellIsRule(operator='equal', formula=['"Modified"'], stopIfTrue=True, fill=modified_fill))
    comparison_sheet.conditional_formatting.add(range_string, CellIsRule(operator='equal', formula=['"Removed"'], stopIfTrue=True, fill=removed_fill))
    
    # Auto-adjust column widths
    column_widths = {
        'A': 20,  # Program
        'B': 30,  # Field
        'C': 20,  # Previous Value
        'D': 20,  # Current Value
        'E': 15   # Status
    }
    
    for col_letter, width in column_widths.items():
        comparison_sheet.column_dimensions[col_letter].width = width
    
    # =========================
    # ADD LEGEND
    # =========================
    # Place legend starting at column G with some spacing
    legend_start_col = 7  # Column G
    legend_start_row = 2  # Start after the header
    
    # Legend title
    legend_title_row = legend_start_row
    legend_cell = comparison_sheet.cell(row=legend_title_row, column=legend_start_col)
    legend_cell.value = "Legend"
    legend_cell.font = Font(bold=True, color="FFFFFF", size=12)
    legend_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    legend_cell.border = thick_border
    legend_cell.alignment = Alignment(horizontal="center", vertical="center")
    comparison_sheet.column_dimensions['G'].width = 15
    
    # Added legend
    added_row = legend_title_row + 1
    added_color_cell = comparison_sheet.cell(row=added_row, column=legend_start_col)
    added_color_cell.value = "Added"
    added_color_cell.fill = added_fill
    added_color_cell.border = thin_border
    added_color_cell.alignment = Alignment(horizontal="center", vertical="center")
    added_color_cell.font = Font(bold=True)
    
    added_desc_cell = comparison_sheet.cell(row=added_row, column=legend_start_col + 1)
    added_desc_cell.value = "New programs added in current version"
    added_desc_cell.border = thin_border
    added_desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    comparison_sheet.column_dimensions['H'].width = 30
    
    # Modified legend
    modified_row = added_row + 1
    modified_color_cell = comparison_sheet.cell(row=modified_row, column=legend_start_col)
    modified_color_cell.value = "Modified"
    modified_color_cell.fill = modified_fill
    modified_color_cell.border = thin_border
    modified_color_cell.alignment = Alignment(horizontal="center", vertical="center")
    modified_color_cell.font = Font(bold=True)
    
    modified_desc_cell = comparison_sheet.cell(row=modified_row, column=legend_start_col + 1)
    modified_desc_cell.value = "Values changed from previous version"
    modified_desc_cell.border = thin_border
    modified_desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Removed legend
    removed_row = modified_row + 1
    removed_color_cell = comparison_sheet.cell(row=removed_row, column=legend_start_col)
    removed_color_cell.value = "Removed"
    removed_color_cell.fill = removed_fill
    removed_color_cell.border = thin_border
    removed_color_cell.alignment = Alignment(horizontal="center", vertical="center")
    removed_color_cell.font = Font(bold=True)
    
    removed_desc_cell = comparison_sheet.cell(row=removed_row, column=legend_start_col + 1)
    removed_desc_cell.value = "Programs no longer in current version"
    removed_desc_cell.border = thin_border
    removed_desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _validate_data(rows: list[dict]) -> list[dict]:
    """
    Validate data to ensure specific properties are met.
    
    Returns a list of validation results with:
    - Category: Type of validation
    - Check: What is being checked
    - Status: Pass or Fail
    - Details: Additional information
    """
    validation_results = []
    
    if not rows:
        validation_results.append({
            "category": "Data Completeness",
            "check": "Data exists",
            "status": "Fail",
            "details": "No data found in extracted rows"
        })
        return validation_results
    
    # Filter out Unspecified programs
    valid_programs = [r for r in rows if r.get("program") and r["program"] != "Unspecified"]
    
    # 1. Data Completeness Checks
    validation_results.append({
        "category": "Data Completeness",
        "check": "Programs with data",
        "status": "Pass" if len(valid_programs) > 0 else "Fail",
        "details": f"Found {len(valid_programs)} programs with valid data"
    })
    
    # 2. Area Data Check
    has_area = sum(1 for r in valid_programs if r.get("PRG_PAR_Area") and float(r["PRG_PAR_Area"]) > 0)
    validation_results.append({
        "category": "Data Completeness",
        "check": "Programs with area data",
        "status": "Pass" if has_area == len(valid_programs) else "Fail",
        "details": f"{has_area}/{len(valid_programs)} programs have area > 0"
    })
    
    # 3. Numerical Threshold - Area > 0
    invalid_areas = [r["program"] for r in valid_programs if not r.get("PRG_PAR_Area") or float(r["PRG_PAR_Area"]) <= 0]
    validation_results.append({
        "category": "Numerical Thresholds",
        "check": "Area > 0 m²",
        "status": "Pass" if len(invalid_areas) == 0 else "Fail",
        "details": f"All areas valid" if len(invalid_areas) == 0 else f"Invalid: {', '.join(invalid_areas[:3])}"
    })
    
    # 4. Use Ratio Check (should be 0-1)
    invalid_ratios = []
    for r in valid_programs:
        ratio = r.get("PRG_PAR_UseRatio")
        if ratio is not None:
            try:
                ratio_val = float(ratio)
                if ratio_val < 0 or ratio_val > 1:
                    invalid_ratios.append(r["program"])
            except (ValueError, TypeError):
                invalid_ratios.append(r["program"])
    
    validation_results.append({
        "category": "Numerical Thresholds",
        "check": "Use Ratio (0-1)",
        "status": "Pass" if len(invalid_ratios) == 0 else "Fail",
        "details": f"All ratios valid" if len(invalid_ratios) == 0 else f"Invalid: {', '.join(invalid_ratios[:3])}"
    })
    
    # 5. Resource Consumption Ratio Check
    invalid_resource_ratios = []
    for r in valid_programs:
        ratio = r.get("PRG_PAR_ResourceConsRatio")
        if ratio is not None:
            try:
                ratio_val = float(ratio)
                if ratio_val < 0 or ratio_val > 1:
                    invalid_resource_ratios.append(r["program"])
            except (ValueError, TypeError):
                invalid_resource_ratios.append(r["program"])
    
    validation_results.append({
        "category": "Numerical Thresholds",
        "check": "Resource Ratio (0-1)",
        "status": "Pass" if len(invalid_resource_ratios) == 0 else "Fail",
        "details": f"All ratios valid" if len(invalid_resource_ratios) == 0 else f"Invalid: {', '.join(invalid_resource_ratios[:3])}"
    })
    
    # 6. Geometry Weight Check (should be positive)
    invalid_weights = []
    for r in valid_programs:
        weight = r.get("PRG_PAR_GeometryWeight")
        if weight is not None:
            try:
                weight_val = float(weight)
                if weight_val < 0:
                    invalid_weights.append(r["program"])
            except (ValueError, TypeError):
                invalid_weights.append(r["program"])
    
    validation_results.append({
        "category": "Numerical Thresholds",
        "check": "Geometry Weight >= 0",
        "status": "Pass" if len(invalid_weights) == 0 else "Fail",
        "details": f"All weights valid" if len(invalid_weights) == 0 else f"Invalid: {', '.join(invalid_weights[:3])}"
    })
    
    # 7. Distance Checks (should be non-negative)
    invalid_mean_dist = []
    invalid_ideal_dist = []
    for r in valid_programs:
        mean_dist = r.get("PRG_PAR_MeanDistToExit")
        if mean_dist is not None:
            try:
                dist_val = float(mean_dist)
                if dist_val < 0:
                    invalid_mean_dist.append(r["program"])
            except (ValueError, TypeError):
                invalid_mean_dist.append(r["program"])
        
        ideal_dist = r.get("PRG_PAR_IdealDistToExit")
        if ideal_dist is not None:
            try:
                dist_val = float(ideal_dist)
                if dist_val < 0:
                    invalid_ideal_dist.append(r["program"])
            except (ValueError, TypeError):
                invalid_ideal_dist.append(r["program"])
    
    validation_results.append({
        "category": "Numerical Thresholds",
        "check": "Mean Distance to Exit >= 0",
        "status": "Pass" if len(invalid_mean_dist) == 0 else "Fail",
        "details": f"All distances valid" if len(invalid_mean_dist) == 0 else f"Invalid: {', '.join(invalid_mean_dist[:3])}"
    })
    
    validation_results.append({
        "category": "Numerical Thresholds",
        "check": "Ideal Distance to Exit >= 0",
        "status": "Pass" if len(invalid_ideal_dist) == 0 else "Fail",
        "details": f"All distances valid" if len(invalid_ideal_dist) == 0 else f"Invalid: {', '.join(invalid_ideal_dist[:3])}"
    })
    
    # 8. Data Consistency Check
    has_all_kpi = sum(1 for r in valid_programs if all([
        r.get("PRG_PAR_Area"),
        r.get("PRG_PAR_UseRatio"),
        r.get("PRG_PAR_ResourceConsRatio"),
        r.get("PRG_PAR_GeometryWeight"),
        r.get("PRG_PAR_MeanDistToExit"),
        r.get("PRG_PAR_IdealDistToExit")
    ]))
    
    validation_results.append({
        "category": "Data Consistency",
        "check": "All KPI parameters present",
        "status": "Pass" if has_all_kpi == len(valid_programs) else "Fail",
        "details": f"{has_all_kpi}/{len(valid_programs)} programs have all KPI parameters"
    })
    
    return validation_results

# =============================================================================
# SECTION 3: EXCEL GENERATION
# =============================================================================


def _create_validation_sheet(workbook: Workbook, rows: list[dict]):
    """
    Create a 'Data Validation' sheet showing validation results.
    """
    validation_data = _validate_data(rows)
    
    if not validation_data:
        return
    
    validation_sheet = workbook.create_sheet("Data Validation")
    
    # Define borders
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Category", "Check", "Status", "Details"]
    validation_sheet.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in validation_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
    
    # Color fills for status
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    
    # Add data rows
    for val in validation_data:
        row = [
            val["category"],
            val["check"],
            val["status"],
            val["details"]
        ]
        validation_sheet.append(row)
        
        # Apply highlighting based on status
        row_idx = validation_sheet.max_row
        status_fill = pass_fill if val["status"] == "Pass" else fail_fill
        
        for col_idx in range(1, 5):
            cell = validation_sheet.cell(row=row_idx, column=col_idx)
            if col_idx == 3:  # Status column
                cell.fill = status_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Auto-adjust column widths
    column_widths = {
        'A': 20,  # Category
        'B': 30,  # Check
        'C': 12,  # Status
        'D': 40   # Details
    }
    
    for col_letter, width in column_widths.items():
        validation_sheet.column_dimensions[col_letter].width = width

    # =========================
    # ADD LEGEND
    # =========================
    legend_start_col = 7  # Column G
    legend_start_row = 2  # Start after the header
    
    # Legend title
    legend_title_row = legend_start_row
    legend_cell = validation_sheet.cell(row=legend_title_row, column=legend_start_col)
    legend_cell.value = "Legend"
    legend_cell.font = Font(bold=True, color="FFFFFF", size=12)
    legend_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    legend_cell.border = thick_border
    legend_cell.alignment = Alignment(horizontal="center", vertical="center")
    validation_sheet.column_dimensions['G'].width = 12
    
    # Pass legend
    pass_row = legend_title_row + 1
    pass_color_cell = validation_sheet.cell(row=pass_row, column=legend_start_col)
    pass_color_cell.value = "Pass"
    pass_color_cell.fill = pass_fill
    pass_color_cell.border = thin_border
    pass_color_cell.alignment = Alignment(horizontal="center", vertical="center")
    pass_color_cell.font = Font(bold=True)
    
    pass_desc_cell = validation_sheet.cell(row=pass_row, column=legend_start_col + 1)
    pass_desc_cell.value = "Validation check passed - data meets requirements"
    pass_desc_cell.border = thin_border
    pass_desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    validation_sheet.column_dimensions['H'].width = 35
    
    # Fail legend
    fail_row = pass_row + 1
    fail_color_cell = validation_sheet.cell(row=fail_row, column=legend_start_col)
    fail_color_cell.value = "Fail"
    fail_color_cell.fill = fail_fill
    fail_color_cell.border = thin_border
    fail_color_cell.alignment = Alignment(horizontal="center", vertical="center")
    fail_color_cell.font = Font(bold=True)
    
    fail_desc_cell = validation_sheet.cell(row=fail_row, column=legend_start_col + 1)
    fail_desc_cell.value = "Validation check failed - data needs review and correction"
    fail_desc_cell.border = thin_border
    fail_desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_excel(rows: list[dict], previous_rows: list[dict] = None) -> str:
    """
    Create an Excel report containing raw data, summary statistics, version comparison, and data validation.
    
    Parameters:
    - rows: Current extracted data rows
    - previous_rows: Optional previous version data for comparison
    
    Process:
    1. Creates a new Excel Workbook.
    2. Sheet 1 ('Program_KPI Parameters'): Writes the raw list of extracted rows.
    3. Sheet 2 ('Summary'): Calculates and writes pivot tables/aggregations (e.g., Area per Tower).
    4. Sheet 3 ('Version Comparison'): If previous_rows provided, shows all changes between versions.
    5. Sheet 4 ('Data Validation'): Shows validation results for data quality checks.
    """
    # Normalize rows to ensure they have PRG_PAR keys for validation/comparison
    _normalize_rows(rows)
    workbook = Workbook()

    # =========================
    # SHEET 1: RAW DATA
    # =========================
    raw_sheet = workbook.active
    raw_sheet.title = "Program_KPI Parameters"

    headers = [
        "program",
        "PRG_PAR_Area",
        "PRG_PAR_UseRatio",
        "PRG_PAR_ResourceConsRatio",
        "PRG_PAR_GeometryWeight",
        "PRG_PAR_MeanDistToExit",
        "PRG_PAR_IdealDistToExit",
    ]
    raw_sheet.append(headers)

    # Define borders
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    light_blue_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")

    # Styling Headers (Bold, Blue Background, Centered)
    header_font = Font(bold=True, color="FFFFFF", size=12)  # Size 12
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in raw_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thick_border

    # Get processed data
    table_rows, grand_totals = _prepare_raw_data(rows)
    previous_program = None

    for row_data in table_rows:
        program = row_data[0]
        # Add light blue gap when program changes
        if previous_program is not None and program != previous_program:
            gap_row = raw_sheet.append([])
            for col in range(1, 8):
                cell = raw_sheet.cell(row=raw_sheet.max_row, column=col)
                cell.fill = light_blue_fill
                cell.border = thin_border
        
        raw_sheet.append(row_data)
        
        # Add thin borders to data row
        data_row = raw_sheet.max_row
        for col in range(1, 8):
            raw_sheet.cell(row=data_row, column=col).border = thin_border
        
        previous_program = program

    # Append Total Row
    total_row = ["Total", grand_totals["area"], "", "", grand_totals["geometry_weight"], "", ""]
    raw_sheet.append(total_row)

    # Style the Total Row (Bold with thick borders)
    last_row = raw_sheet.max_row
    for col in range(1, 8):
        cell = raw_sheet.cell(row=last_row, column=col)
        cell.font = Font(bold=True, size=12)
        cell.border = thick_border

    # Auto-adjust column widths
    for col_idx, column_cells in enumerate(raw_sheet.columns, start=1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        raw_sheet.column_dimensions[column].width = adjusted_width

    # =========================
    # DATA AGGREGATION
    # =========================
    summary_data_dict = _prepare_summary_data(rows)
    matrix = summary_data_dict["matrix"]

    # =========================
    # SHEET 2: SUMMARY
    # =========================
    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Metric", "Value"])
    
    # Style first header row
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.border = thick_border
        cell.alignment = Alignment(horizontal="center")
    
    # Total Area row
    total_row_idx = summary_sheet.max_row + 1
    summary_sheet.append(["Total Area (m2)", summary_data_dict["total_area"]])
    for col in range(1, 3):
        summary_sheet.cell(row=total_row_idx, column=col).border = thin_border
    summary_sheet.append([])

    summary_sheet.append(["Area per Tower"])
    header_row = summary_sheet.max_row
    for cell in summary_sheet[header_row]:
        if cell.value:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.border = thick_border
    
    summary_sheet.append(["Tower", "Area (m2)"])
    header_row = summary_sheet.max_row
    for col in range(1, 3):
        cell = summary_sheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.border = thick_border
    
    for tower, area in sorted(summary_data_dict["area_per_tower"].items()):
        summary_sheet.append([tower, area])
        data_row = summary_sheet.max_row
        for col in range(1, 3):
            summary_sheet.cell(row=data_row, column=col).border = thin_border
    summary_sheet.append([])

    summary_sheet.append(["Area per Level"])
    header_row = summary_sheet.max_row
    for cell in summary_sheet[header_row]:
        if cell.value:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.border = thick_border
    
    summary_sheet.append(["Level", "Area (m2)"])
    header_row = summary_sheet.max_row
    for col in range(1, 3):
        cell = summary_sheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.border = thick_border
    
    for level, area in sorted(summary_data_dict["area_per_level"].items()):
        summary_sheet.append([level, area])
        data_row = summary_sheet.max_row
        for col in range(1, 3):
            summary_sheet.cell(row=data_row, column=col).border = thin_border
    summary_sheet.append([])

    summary_sheet.append(["Area per Program"])
    header_row = summary_sheet.max_row
    for cell in summary_sheet[header_row]:
        if cell.value:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.border = thick_border
    
    summary_sheet.append(["Program", "Area (m2)"])
    header_row = summary_sheet.max_row
    for col in range(1, 3):
        cell = summary_sheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.border = thick_border
    
    for program, area in sorted(summary_data_dict["area_per_program"].items()):
        summary_sheet.append([program, area])
        data_row = summary_sheet.max_row
        for col in range(1, 3):
            summary_sheet.cell(row=data_row, column=col).border = thin_border
    summary_sheet.append([])

    # Create a Matrix: Levels (rows) x Towers (columns)
    towers_sorted = sorted(summary_data_dict["area_per_tower"].keys())
    levels_sorted = sorted(matrix.keys())

    summary_sheet.append(["Tower x Level Matrix (m2)"])
    header_row = summary_sheet.max_row
    for cell in summary_sheet[header_row]:
        if cell.value:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.border = thick_border
    
    summary_sheet.append(["Level"] + towers_sorted)
    header_row = summary_sheet.max_row
    for col in range(1, len(towers_sorted) + 2):
        cell = summary_sheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.border = thick_border

    for level in levels_sorted:
        summary_sheet.append([level] + [matrix[level].get(t, 0.0) for t in towers_sorted])
        data_row = summary_sheet.max_row
        for col in range(1, len(towers_sorted) + 2):
            summary_sheet.cell(row=data_row, column=col).border = thin_border

    # =========================
    # SHEET 3: VERSION COMPARISON (if previous version exists)
    # =========================
    if previous_rows is not None:
        _create_comparison_sheet(workbook, rows, previous_rows)

    # =========================
    # SHEET 4: DATA VALIDATION
    # =========================
    _create_validation_sheet(workbook, rows)

    output_path = os.path.join(os.getcwd(), "capsule_areas.xlsx")
    workbook.save(output_path)

    return output_path


# =============================================================================
# SECTION 4: GOOGLE SHEETS GENERATION
# =============================================================================

def update_google_sheet(
    rows: list[dict], sheet_id: str, service_account_json: str, previous_rows: list[dict] = None
) -> str:
    """
    Authenticate with Google and update a specific Spreadsheet with the report data.
    
    Parameters:
    - rows: Current extracted data rows
    - sheet_id: The ID of the target Google Sheet
    - service_account_json: The JSON credentials for Google API access
    - previous_rows: Optional previous version data for comparison
    
    Process:
    1. Validates inputs (Sheet ID and JSON credentials).
    2. Authenticates using the Service Account credentials.
    3. Connects to the Google Sheet by ID.
    4. Clears and updates the 'Program_KPI Parameters' tab with raw data.
    5. Clears and updates the 'Summary' tab with calculated statistics.
    6. If previous_rows provided, creates 'Version Comparison' tab with changes.
    """
    if not sheet_id:
        raise ValueError("Google Sheet ID is required for this format.")
    if not service_account_json:
        raise ValueError("Service Account JSON is required for this format.")

    try:
        # Clean input: remove whitespace and potential wrapping quotes
        json_str = service_account_json.strip()
        if json_str.startswith("'") and json_str.endswith("'"):
            json_str = json_str[1:-1]

        creds_dict = json.loads(json_str)
        
        if not isinstance(creds_dict, dict):
            raise ValueError("The provided JSON is not a dictionary. Please paste the full content of the JSON file.")
        if "private_key" not in creds_dict or "client_email" not in creds_dict:
            raise ValueError("The JSON key is missing required fields ('private_key' or 'client_email').")
            
        # Define scopes to allow reading/writing to Sheets and Drive
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        # Create credentials object
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        # Authorize the gspread client
        client = gspread.authorize(creds)
        sh = client.open_by_key(sheet_id)
    except JSONDecodeError as e:
        raise ValueError(
            f"Invalid JSON format in Service Account Key. "
            f"It seems the key is truncated or has a syntax error. "
            f"Error details: {e}"
        )
    except Exception as e:
        print("Detailed error traceback:")
        traceback.print_exc()
        raise ValueError(f"Google Sheets Error ({type(e).__name__}): {e}")
    
    # --- Define Styles for Google Sheets (using raw dicts for compatibility) ---
    def rgb(r, g, b):
        return {"red": r, "green": g, "blue": b}

    blue_bg = rgb(0.31, 0.51, 0.74)      # 4F81BD
    white_fg = rgb(1, 1, 1)
    green_bg = rgb(0.78, 0.94, 0.81)     # C6EFCE
    yellow_bg = rgb(1, 0.95, 0.8)        # FFF2CC
    red_bg = rgb(1, 0.78, 0.81)          # FFC7CE
    
    header_fmt = {
        "backgroundColor": blue_bg,
        "textFormat": {"foregroundColor": white_fg, "bold": True},
        "horizontalAlignment": "CENTER"
    }
    border_fmt = {
        "borders": {"top": {"style": "SOLID"}, "bottom": {"style": "SOLID"}, "left": {"style": "SOLID"}, "right": {"style": "SOLID"}}
    }
    bold_fmt = {"textFormat": {"bold": True}}

    # Helper to construct GridRange dict manually (used in Comparison and Validation)
    def get_grid_range(ws, start_row, end_row, start_col, end_col):
        return {
            "sheetId": ws.id,
            "startRowIndex": start_row - 1,
            "endRowIndex": end_row,
            "startColumnIndex": start_col - 1,
            "endColumnIndex": end_col
        }

    # Normalize rows to ensure they have PRG_PAR keys for validation/comparison
    _normalize_rows(rows)

    # 1. Raw Data
    # Select or create the worksheet for raw data
    try:
        ws_raw = sh.worksheet("Program_KPI Parameters")
        
        # --- NEW LOGIC: Fetch previous data from Sheet if not provided (for Automate) ---
        if previous_rows is None:
            try:
                # Check if sheet has data (headers + at least 1 row)
                if len(ws_raw.get_values("A1:A2")) > 1:
                    previous_rows = ws_raw.get_all_records()
                    print(f"✓ Loaded {len(previous_rows)} previous rows from existing Google Sheet.")
            except Exception as e:
                print(f"⚠ Could not read previous data from Sheet: {e}")
        # -------------------------------------------------------------------------------

        ws_raw.clear()
    except gspread.WorksheetNotFound:
        ws_raw = sh.add_worksheet(title="Program_KPI Parameters", rows=1000, cols=20)

    headers = [
        "program",
        "PRG_PAR_Area",
        "PRG_PAR_UseRatio",
        "PRG_PAR_ResourceConsRatio",
        "PRG_PAR_GeometryWeight",
        "PRG_PAR_MeanDistToExit",
        "PRG_PAR_IdealDistToExit",
    ]

    # Get processed data
    table_rows, grand_totals = _prepare_raw_data(rows)
    
    data_values = [headers] + table_rows
    
    # Append Total Row
    data_values.append(["Total", grand_totals["area"], "", "", grand_totals["geometry_weight"], "", ""])
    
    ws_raw.update(values=data_values)

    # Format Total Row in Google Sheets (Bold)
    total_row_idx = len(data_values)
    # Apply bold formatting to the last row
    ws_raw.format(f"A{total_row_idx}:G{total_row_idx}", bold_fmt)
    # Apply Header style
    ws_raw.format("A1:G1", header_fmt)
    # Apply Borders to all data
    ws_raw.format(f"A1:G{total_row_idx}", border_fmt)

    # 2. Summary
    # Calculate aggregates (reusing logic from Excel generation)
    summary_data_dict = _prepare_summary_data(rows)
    
    # Select or create the worksheet for summary data
    try:
        ws_summary = sh.worksheet("Summary")
        ws_summary.clear()
    except gspread.WorksheetNotFound:
        ws_summary = sh.add_worksheet(title="Summary", rows=100, cols=10)

    # Track header rows for formatting
    header_rows = []
    
    summary_data = [
        ["Metric", "Value"],
        ["Total Area (m2)", summary_data_dict["total_area"]],
        [],
        ["Area per Tower"],
        ["Tower", "Area (m2)"],
    ]
    header_rows.extend([1, 4, 5]) # 1-based indices
    
    for tower, area in sorted(summary_data_dict["area_per_tower"].items()):
        summary_data.append([tower, area])
    
    summary_data.append([])
    summary_data.append(["Area per Level"])
    summary_data.append(["Level", "Area (m2)"])
    
    current_row = len(summary_data)
    header_rows.extend([current_row - 1, current_row])
    
    for level, area in sorted(summary_data_dict["area_per_level"].items()):
        summary_data.append([level, area])

    summary_data.append([])
    summary_data.append(["Area per Program"])
    summary_data.append(["Program", "Area (m2)"])
    
    current_row = len(summary_data)
    header_rows.extend([current_row - 1, current_row])
    
    for program, area in sorted(summary_data_dict["area_per_program"].items()):
        summary_data.append([program, area])

    # Matrix Tower x Level aggregation
    summary_data.append([])
    summary_data.append(["Tower x Level Matrix (m2)"])
    
    current_row = len(summary_data)
    header_rows.append(current_row)
    
    towers_sorted = sorted(summary_data_dict["area_per_tower"].keys())
    matrix = summary_data_dict["matrix"]
    levels_sorted = sorted(matrix.keys())
    
    # Header row for matrix
    summary_data.append(["Level"] + towers_sorted)
    header_rows.append(len(summary_data))
    
    # Data rows for matrix
    for level in levels_sorted:
        row = [level] + [matrix[level].get(t, 0.0) for t in towers_sorted]
        summary_data.append(row)

    # Use named argument 'values' for compatibility with newer gspread versions
    ws_summary.update(values=summary_data)
    
    # Apply formatting to Summary
    # Borders for the whole range
    max_rows = len(summary_data)
    max_cols = max(len(row) for row in summary_data) if summary_data else 1
    col_letter = get_column_letter(max_cols)
    ws_summary.format(f"A1:{col_letter}{max_rows}", border_fmt)
    
    # Headers
    for row_idx in header_rows:
        # Determine width of this specific header row
        row_len = len(summary_data[row_idx-1])
        if row_len > 0:
            letter = get_column_letter(row_len)
            ws_summary.format(f"A{row_idx}:{letter}{row_idx}", header_fmt)

    # Helper to construct GridRange dict manually (used in Comparison and Validation)
    def get_grid_range(ws, start_row, end_row, start_col, end_col):
        return {
            "sheetId": ws.id,
            "startRowIndex": start_row - 1,
            "endRowIndex": end_row,
            "startColumnIndex": start_col - 1,
            "endColumnIndex": end_col
        }

    # 3. Version Comparison (if previous version exists)
    if previous_rows is not None and len(previous_rows) > 0:
        comparison_data = _compare_versions(rows, previous_rows)
        
        try:
            ws_comparison = sh.worksheet("Version Comparison")
            ws_comparison.clear()
        except gspread.WorksheetNotFound:
            ws_comparison = sh.add_worksheet(title="Version Comparison", rows=1000, cols=10)
        
        comparison_values = [["Program", "Field", "Previous Value", "Current Value", "Status"]]
        
        for comp in comparison_data:
            comparison_values.append([
                comp["program"],
                comp["field"],
                comp["previous_value"],
                comp["current_value"],
                comp["status"]
            ])
        
        ws_comparison.update(values=comparison_values)
        
        # Formatting
        row_count = len(comparison_values)
        ws_comparison.format("A1:E1", header_fmt)
        ws_comparison.format(f"A1:E{row_count}", border_fmt)
        
        # Conditional Formatting for Status (Column E)
        # Added -> Green, Modified -> Yellow, Removed -> Red
        
        # Construct batch update requests for conditional formatting
        requests = []
        status_col_idx = 5 # Column E
        
        conditions = [
            ("Added", green_bg),
            ("Modified", yellow_bg),
            ("Removed", red_bg)
        ]
        
        for status_text, bg_color in conditions:
            requests.append({
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [get_grid_range(ws_comparison, 2, row_count, status_col_idx, status_col_idx)],
                        "booleanRule": {
                            "condition": {"type": "TEXT_EQ", "values": [{"userEnteredValue": status_text}]},
                            "format": {"backgroundColor": bg_color}
                        }
                    },
                    "index": 0
                }
            })
            
        if requests:
            sh.batch_update({"requests": requests})

    # 4. Data Validation
    validation_data = _validate_data(rows)
    
    try:
        ws_validation = sh.worksheet("Data Validation")
        ws_validation.clear()
    except gspread.WorksheetNotFound:
        ws_validation = sh.add_worksheet(title="Data Validation", rows=100, cols=10)

    validation_headers = ["Category", "Check", "Status", "Details"]
    validation_values = [validation_headers]
    
    for val in validation_data:
        validation_values.append([
            val["category"],
            val["check"],
            val["status"],
            val["details"]
        ])
    
    ws_validation.update(values=validation_values)
    
    # Formatting
    row_count = len(validation_values)
    ws_validation.format("A1:D1", header_fmt)
    ws_validation.format(f"A1:D{row_count}", border_fmt)
    
    # Conditional Formatting for Status (Column C)
    # Pass -> Green, Fail -> Red
    requests = []
    status_col_idx = 3 # Column C
    
    conditions = [
        ("Pass", green_bg),
        ("Fail", red_bg)
    ]
    
    for status_text, bg_color in conditions:
        requests.append({
            "addConditionalFormatRule": {
                "rule": {
                    "ranges": [get_grid_range(ws_validation, 2, row_count, status_col_idx, status_col_idx)],
                    "booleanRule": {
                        "condition": {"type": "TEXT_EQ", "values": [{"userEnteredValue": status_text}]},
                        "format": {"backgroundColor": bg_color}
                    }
                },
                "index": 0
            }
        })

    if requests:
        sh.batch_update({"requests": requests})

    return sh.url


# =============================================================================
# SECTION 5: UTILS & TESTING
# =============================================================================

def load_previous_data_from_file(file_path: str) -> list[dict]:
    """
    Load previous version data from a saved Excel file.
    
    This is useful for comparing against an exported baseline.
    """
    try:
        workbook = load_workbook(file_path)
    except Exception as e:
        print(f"⚠ Could not load previous file: {e}")
        return []

    if "Program_KPI Parameters" not in workbook.sheetnames:
        print("⚠ Sheet 'Program_KPI Parameters' not found in previous file.")
        return []

    sheet = workbook["Program_KPI Parameters"]
    
    rows = []
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            # Skip header row
            continue
        elif row[0] and row[0] != "Total":  # Skip empty and total rows
            row_dict = {
                "program": row[0],
                "PRG_PAR_Area": row[1],
                "PRG_PAR_UseRatio": row[2],
                "PRG_PAR_ResourceConsRatio": row[3],
                "PRG_PAR_GeometryWeight": row[4],
                "PRG_PAR_MeanDistToExit": row[5],
                "PRG_PAR_IdealDistToExit": row[6],
            }
            rows.append(row_dict)
    
    print(f"✓ Loaded {len(rows)} programs from {file_path}")
    return rows


def _create_test_data():
    """Create sample test data with some validation issues (for local testing)."""
    return [
        {
            "tower": "Tower A", "level": "L1", "capsule": "CAP001", "program": "Office A", "location": "North Wing",
            "area": 500.5, "use_ratio": 0.8, "resource_cons_ratio": 0.6, "geometry_weight": 1.5, "mean_dist_to_exit": 15.0, "ideal_dist_to_exit": 10.0,
            "PRG_PAR_Area": 500.5, "PRG_PAR_UseRatio": 0.8, "PRG_PAR_ResourceConsRatio": 0.6, "PRG_PAR_GeometryWeight": 1.5, "PRG_PAR_MeanDistToExit": 15.0, "PRG_PAR_IdealDistToExit": 10.0,
        },
        {
            "tower": "Tower A", "level": "L2", "capsule": "CAP002", "program": "Office B", "location": "North Wing",
            "area": 0, "use_ratio": 1.5, "resource_cons_ratio": 0.5, "geometry_weight": 1.2, "mean_dist_to_exit": 20.0, "ideal_dist_to_exit": 12.0,
            "PRG_PAR_Area": 0, "PRG_PAR_UseRatio": 1.5, "PRG_PAR_ResourceConsRatio": 0.5, "PRG_PAR_GeometryWeight": 1.2, "PRG_PAR_MeanDistToExit": 20.0, "PRG_PAR_IdealDistToExit": 12.0,
        },
        {
            "tower": "Tower B", "level": "L1", "capsule": "CAP003", "program": "Warehouse C", "location": "South Wing",
            "area": 2000.0, "use_ratio": 0.9, "resource_cons_ratio": 0.7, "geometry_weight": 2.0, "mean_dist_to_exit": 50.0, "ideal_dist_to_exit": 40.0,
            "PRG_PAR_Area": 2000.0, "PRG_PAR_UseRatio": 0.9, "PRG_PAR_ResourceConsRatio": 0.7, "PRG_PAR_GeometryWeight": 2.0, "PRG_PAR_MeanDistToExit": 50.0, "PRG_PAR_IdealDistToExit": 40.0,
        },
        {
            "tower": "Tower C", "level": "L3", "capsule": "CAP004", "program": "Retail D", "location": "East Wing",
            "area": 750.0, "use_ratio": 0.6, "resource_cons_ratio": -0.1, "geometry_weight": -0.5, "mean_dist_to_exit": -10.0, "ideal_dist_to_exit": 20.0,
            "PRG_PAR_Area": 750.0, "PRG_PAR_UseRatio": 0.6, "PRG_PAR_ResourceConsRatio": -0.1, "PRG_PAR_GeometryWeight": -0.5, "PRG_PAR_MeanDistToExit": -10.0, "PRG_PAR_IdealDistToExit": 20.0,
        },
    ]


if __name__ == "__main__":
    print("=" * 60)
    print("DATA VALIDATION FEATURE TEST (Local)")
    print("=" * 60)
    
    # Create test data
    test_rows = _create_test_data()
    print(f"\nCreated test data with {len(test_rows)} programs")
    
    # Run validation
    print("\n" + "=" * 60)
    print("RUNNING VALIDATION CHECKS")
    print("=" * 60)
    
    validation_results = _validate_data(test_rows)
    
    print(f"\nValidation Results ({len(validation_results)} checks):")
    print("-" * 60)
    
    current_category = None
    for result in validation_results:
        if result["category"] != current_category:
            print(f"\n[{result['category']}]")
            current_category = result["category"]
        
        status_mark = "PASS" if result["status"] == "Pass" else "FAIL"
        print(f"  {status_mark:4} | {result['check']:30} | {result['details']}")
    
    # Generate Excel with validation sheet
    print("\n" + "=" * 60)
    print("GENERATING EXCEL REPORT")
    print("=" * 60)
    
    try:
        output_path = generate_excel(test_rows)
        print(f"\nExcel report generated: {output_path}")
        
        # Verify the Data Validation sheet exists
        workbook = load_workbook(output_path)
        sheet_names = workbook.sheetnames
        print(f"\nWorkbook sheets created:")
        for idx, sheet_name in enumerate(sheet_names, 1):
            print(f"  {idx}. {sheet_name}")
            
            # If it's the Data Validation sheet, show preview
            if sheet_name == "Data Validation":
                ws = workbook[sheet_name]
                print(f"     [Data Validation sheet with {ws.max_row - 1} validation checks]")
                
                # Show first few rows
                print(f"     Sample results:")
                for row_idx in range(2, min(5, ws.max_row + 1)):
                    cells = [ws.cell(row_idx, col).value for col in range(1, 5)]
                    if cells[2] == "Pass":
                        print(f"       [PASS] {cells[1]}: {cells[3]}")
                    else:
                        print(f"       [FAIL] {cells[1]}: {cells[3]}")
        
        print("\n" + "=" * 60)
        print("SUCCESS: Data Validation feature is working correctly!")
        print("=" * 60)
        
    except Exception as e:
        print(f"\nError generating Excel report: {e}")
        traceback.print_exc()