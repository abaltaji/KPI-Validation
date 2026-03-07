"""
Module for comparing previous and current versions of KPI data.

This module provides utilities for loading baseline data, generating comparison reports,
and highlighting changes between versions.
"""

import importlib
from pathlib import Path

# Import modules using importlib because they start with numbers
extraction_module = importlib.import_module("03_extraction")
extract_capsule_areas = extraction_module.extract_capsule_areas

reporting_module = importlib.import_module("04_reporting")
generate_excel = reporting_module.generate_excel


def generate_comparison_report(current_model, previous_model):
    """
    Generate an Excel report comparing current and previous model versions.
    
    Parameters:
    - current_model: The current Speckle model
    - previous_model: The previous version of the Speckle model
    
    Returns:
    - output_path: Path to the generated Excel file
    
    The generated Excel file will include:
    1. Program_KPI Parameters sheet: Current data
    2. Summary sheet: Summary statistics for current data
    3. Version Comparison sheet: All changes between versions (highlighted)
    """
    # Extract data from both versions
    current_rows = extract_capsule_areas(current_model)
    previous_rows = extract_capsule_areas(previous_model)
    
    # Generate Excel with comparison
    output_path = generate_excel(current_rows, previous_rows=previous_rows)
    
    print(f"✓ Comparison report generated: {output_path}")
    return output_path


def generate_report_without_comparison(current_model):
    """
    Generate a standard Excel report without version comparison.
    
    Parameters:
    - current_model: The current Speckle model
    
    Returns:
    - output_path: Path to the generated Excel file
    
    The generated Excel file will include:
    1. Program_KPI Parameters sheet: Current data
    2. Summary sheet: Summary statistics
    """
    current_rows = extract_capsule_areas(current_model)
    
    # Generate Excel without comparison (no previous_rows parameter)
    output_path = generate_excel(current_rows)
    
    print(f"✓ Report generated: {output_path}")
    return output_path


def load_previous_data_from_file(file_path):
    """
    Load previous version data from a saved Excel file.
    
    This is useful for comparing against an exported baseline.
    Note: This requires openpyxl package.
    
    Parameters:
    - file_path: Path to the Excel file containing previous data
    
    Returns:
    - rows: List of row dictionaries with the previous data
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required to load data from Excel files")
    
    workbook = load_workbook(file_path)
    sheet = workbook["Program_KPI Parameters"]
    
    rows = []
    headers = None
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            # Extract headers (skip the Total row)
            headers = [h.lower().replace(" ", "_") if h else f"col_{i}" 
                      for i, h in enumerate(row, start=1)]
        elif row[0] and row[0] != "Total":  # Skip empty and total rows
            row_dict = {
                "program": row[0],
                "PRG_PAR_Area": row[1],
                "PRG_PAR_UseRatio": row[2],
                "PRG_PAR_ResourceConsRatio": row[3],
                "PRG_PAR_GeometryWeight": row[4],
                "PRG_PAR_MeanDistToExit": row[5],
                "PRG_PAR_IdealDistToExit": row[6],
            }
            rows.append(row_dict)
    
    print(f"✓ Loaded {len(rows)} programs from {file_path}")
    return rows


if __name__ == "__main__":
    print("""
    Version Comparison Module (07_version_comparison)
    ==================================================
    
    This module provides utilities for comparing KPI data between versions.
    
    Key Functions:
    1. generate_comparison_report(current_model, previous_model)
       - Creates Excel with Version Comparison sheet showing all changes
       - Color-coded cells (green=Added, yellow=Modified, red=Removed)
    
    2. generate_report_without_comparison(current_model)
       - Standard report without version comparison
    
    3. load_previous_data_from_file(file_path)
       - Load previous version data from a saved Excel file
       - Useful for comparing against a baseline export
    
    Changes Tracked:
    - Added programs (new to current version)
    - Modified values (any numeric difference detected)
    - Removed programs (no longer in current version)
    - Field-level comparison with Previous/Current/Status columns
    
    Usage:
    ------
    from 07_version_comparison import load_previous_data_from_file, generate_comparison_report
    
    previous_rows = load_previous_data_from_file("baseline_2025.xlsx")
    current_rows = extract_capsule_areas(current_model)
    output = generate_excel(current_rows, previous_rows=previous_rows)
    """)
