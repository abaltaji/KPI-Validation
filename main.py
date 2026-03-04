"""Module for Speckle Automate KPI validation.

Authentication module for Speckle - provides a reusable get_client() function.
"""

import os
import json
import traceback
from json import JSONDecodeError
from collections import defaultdict
from typing import Any

from dotenv import load_dotenv
# INSTRUCTION: Uncommented openpyxl. You MUST have openpyxl installed 
# in your environment/requirements.txt or the Workbook() call will fail.
from openpyxl import Workbook 
# INSTRUCTION: You MUST have 'gspread' and 'google-auth' installed.
import gspread
from google.oauth2.service_account import Credentials
from speckle_automate import (
    AutomationContext,
    execute_automate_function,
)
from specklepy.api.client import SpeckleClient

from flatten import flatten_base
from function_inputs import FunctionInputs, OutputFormat


def get_client() -> SpeckleClient:
    """
    Authenticate and return a SpeckleClient instance.
    
    Requires SPECKLE_TOKEN in environment or .env file.
    Optionally set SPECKLE_SERVER (defaults to app.speckle.systems).
    """
    load_dotenv()

    token = os.environ.get("SPECKLE_TOKEN")
    server_host = os.environ.get("SPECKLE_SERVER", "app.speckle.systems")

    if not token:
        raise ValueError("Set SPECKLE_TOKEN in your .env file and re-run.")

    client = SpeckleClient(host=server_host)
    client.authenticate_with_token(token)

    return client


def upload_file_to_speckle(
    client: SpeckleClient, project_id: str, file_path: str, file_name: str, token: str
) -> str:
    """Upload a file to Speckle using the REST API."""
    import requests
    
    # INSTRUCTION: Use the 'token' argument passed to the function.
    # Do NOT use os.environ.get here as it fails in the cloud.
    url = f"{client.url}/api/file/create"
    
    with open(file_path, "rb") as f:
        files = {"files": (file_name, f)}
        params = {"streamId": project_id}
        headers = {"Authorization": f"Bearer {token}"} # Fixed to use the argument
        
        response = requests.post(url, files=files, params=params, headers=headers)
        response.raise_for_status()
        
        result = response.json()
        file_id = result.get("fileIds", [None])[0]
        
        if not file_id:
            raise ValueError("No file ID returned from upload")
        
        return file_id


def post_comment_with_file(
    client: SpeckleClient, model_id: str, project_id: str, file_id: str, file_name: str
) -> None:
    """Post a comment on the model with the uploaded file."""
    comment_text = f"📊 KPI Validation Report: {file_name}"
    client.comment.create(
        stream_id=project_id,
        object_id=model_id,
        text=comment_text,
        resources=[{"resourceType": "file", "resourceId": file_id}],
    )


def _get_attr(obj: Any, *names: str, default=None):
    """Safely get first available attribute/key from object or dict."""
    for name in names:
        if isinstance(obj, dict) and name in obj:
            return obj[name]
        if hasattr(obj, name):
            return getattr(obj, name)
    return default


def extract_capsule_areas(model: Any) -> list[dict]:
    """Extract area rows from a Speckle model."""
    rows: list[dict] = []
    for item in flatten_base(model):
        # INSTRUCTION: For Meshes, ensure they have a custom attribute called "area".
        # Standard Speckle Meshes do not calculate area automatically in Automate.
        area = _get_attr(item, "area", "Area", default=None)
        if area is None:
            continue

        try:
            area_value = float(area)
        except (TypeError, ValueError):
            continue

        rows.append(
            {
                "tower": _get_attr(item, "tower", "Tower", default="Unknown"),
                "level": _get_attr(item, "level", "Level", default="Unspecified"),
                "capsule": _get_attr(
                    item, "capsule", "capsule_no", "CapsuleNo", default=""
                ),
                "program": _get_attr(item, "program", "Program", default="Unspecified"),
                "location": _get_attr(item, "location", "Location", default=""),
                "area": area_value,
            }
        )
    return rows


def generate_excel(rows: list[dict]) -> str:
    """Generate an Excel file from the rows."""
    workbook = Workbook()

    # =========================
    # SHEET 1: RAW DATA
    # =========================
    raw_sheet = workbook.active
    raw_sheet.title = "Capsule Areas"

    headers = ["Tower", "Level", "Capsule No.", "Program", "Location", "2D Area (m2)"]
    raw_sheet.append(headers)

    for r in rows:
        raw_sheet.append(
            [r["tower"], r["level"], r["capsule"], r["program"], r["location"], r["area"]]
        )

    # =========================
    # DATA AGGREGATION
    # =========================
    total_area = sum(r["area"] for r in rows)

    area_per_tower: dict[str, float] = defaultdict(float)
    area_per_level: dict[str, float] = defaultdict(float)
    area_per_program: dict[str, float] = defaultdict(float)
    matrix: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))

    for r in rows:
        tower = r["tower"] or "Unknown"
        level = r["level"] or "Unspecified"
        program = r["program"] or "Unspecified"

        area_per_tower[tower] += r["area"]
        area_per_level[level] += r["area"]
        area_per_program[program] += r["area"]
        matrix[level][tower] += r["area"]

    # =========================
    # SHEET 2: SUMMARY
    # =========================
    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Total Area (m2)", total_area])
    summary_sheet.append([])

    summary_sheet.append(["Area per Tower"])
    summary_sheet.append(["Tower", "Area (m2)"])
    for tower, area in sorted(area_per_tower.items()):
        summary_sheet.append([tower, area])
    summary_sheet.append([])

    summary_sheet.append(["Area per Level"])
    summary_sheet.append(["Level", "Area (m2)"])
    for level, area in sorted(area_per_level.items()):
        summary_sheet.append([level, area])
    summary_sheet.append([])

    summary_sheet.append(["Area per Program"])
    summary_sheet.append(["Program", "Area (m2)"])
    for program, area in sorted(area_per_program.items()):
        summary_sheet.append([program, area])
    summary_sheet.append([])

    # Tower x Level matrix
    towers_sorted = sorted(area_per_tower.keys())
    levels_sorted = sorted(matrix.keys())

    summary_sheet.append(["Tower x Level Matrix (m2)"])
    summary_sheet.append(["Level"] + towers_sorted)

    for level in levels_sorted:
        summary_sheet.append([level] + [matrix[level].get(t, 0.0) for t in towers_sorted])

    output_path = "capsule_areas.xlsx"
    workbook.save(output_path)

    return output_path


def update_google_sheet(
    rows: list[dict], sheet_id: str, service_account_json: str
) -> str:
    """Update a Google Sheet with the rows and summary."""
    if not sheet_id:
        raise ValueError("Google Sheet ID is required for this format.")
    if not service_account_json:
        raise ValueError("Service Account JSON is required for this format.")

    # Authenticate
    try:
        # Clean input: remove whitespace and potential wrapping quotes
        json_str = service_account_json.strip()
        if json_str.startswith("'") and json_str.endswith("'"):
            json_str = json_str[1:-1]

        creds_dict = json.loads(json_str)
        
        if not isinstance(creds_dict, dict):
            raise ValueError("The provided JSON is not a dictionary. Please paste the full content of the JSON file.")
        if "private_key" not in creds_dict or "client_email" not in creds_dict:
            raise ValueError("The JSON key is missing required fields ('private_key' or 'client_email').")
            
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        client = gspread.authorize(creds)
        sh = client.open_by_key(sheet_id)
    except JSONDecodeError as e:
        raise ValueError(
            f"Invalid JSON format in Service Account Key. "
            f"It seems the key is truncated or has a syntax error. "
            f"Error details: {e}"
        )
    except Exception as e:
        print("Detailed error traceback:")
        traceback.print_exc()
        raise ValueError(f"Google Sheets Error ({type(e).__name__}): {e}")

    # 1. Raw Data
    try:
        ws_raw = sh.worksheet("Capsule Areas")
        ws_raw.clear()
    except gspread.WorksheetNotFound:
        ws_raw = sh.add_worksheet(title="Capsule Areas", rows=1000, cols=20)

    headers = ["Tower", "Level", "Capsule No.", "Program", "Location", "2D Area (m2)"]
    data_values = [headers] + [
        [r["tower"], r["level"], r["capsule"], r["program"], r["location"], r["area"]]
        for r in rows
    ]
    ws_raw.update(data_values)

    # 2. Summary
    # Calculate aggregates (reusing logic)
    total_area = sum(r["area"] for r in rows)
    
    area_per_tower = defaultdict(float)
    area_per_level = defaultdict(float)
    area_per_program = defaultdict(float)
    matrix = defaultdict(lambda: defaultdict(float))

    for r in rows:
        tower = r["tower"] or "Unknown"
        level = r["level"] or "Unspecified"
        program = r["program"] or "Unspecified"

        area_per_tower[tower] += r["area"]
        area_per_level[level] += r["area"]
        area_per_program[program] += r["area"]
        matrix[level][tower] += r["area"]

    try:
        ws_summary = sh.worksheet("Summary")
        ws_summary.clear()
    except gspread.WorksheetNotFound:
        ws_summary = sh.add_worksheet(title="Summary", rows=100, cols=10)

    summary_data = [
        ["Metric", "Value"],
        ["Total Area (m2)", total_area],
        [],
        ["Area per Tower"],
        ["Tower", "Area (m2)"],
    ]
    for tower, area in sorted(area_per_tower.items()):
        summary_data.append([tower, area])
    
    summary_data.append([])
    summary_data.append(["Area per Level"])
    summary_data.append(["Level", "Area (m2)"])
    for level, area in sorted(area_per_level.items()):
        summary_data.append([level, area])

    summary_data.append([])
    summary_data.append(["Area per Program"])
    summary_data.append(["Program", "Area (m2)"])
    for program, area in sorted(area_per_program.items()):
        summary_data.append([program, area])

    # Matrix Tower x Level
    summary_data.append([])
    summary_data.append(["Tower x Level Matrix (m2)"])
    
    towers_sorted = sorted(area_per_tower.keys())
    levels_sorted = sorted(matrix.keys())
    
    # Header row for matrix
    summary_data.append(["Level"] + towers_sorted)
    
    # Data rows for matrix
    for level in levels_sorted:
        row = [level] + [matrix[level].get(t, 0.0) for t in towers_sorted]
        summary_data.append(row)

    # Use named argument 'values' for compatibility with newer gspread versions
    ws_summary.update(values=summary_data)

    return sh.url


def automate_function(
    automation_context: AutomationContext, function_inputs: FunctionInputs
) -> None:
    """Simplified entry point using the official SDK file storage."""
    
    # 1. Receive the model version data
    version_root_object = automation_context.receive_version()
    
    if not version_root_object:
        automation_context.mark_run_failed("No model data received.")
        return
    
    # 2. Extract Data
    rows = extract_capsule_areas(version_root_object)
    
    if not rows:
        automation_context.mark_run_success("No 2D area data found in the model meshes.")
        return

    # 3. Process based on Output Format
    try:
        if function_inputs.output_format == OutputFormat.GOOGLE_SHEET:
            sheet_url = update_google_sheet(
                rows,
                function_inputs.google_sheet_id,
                function_inputs.google_service_account_json.get_secret_value(),
            )
            automation_context.mark_run_success(f"✓ Google Sheet updated: {sheet_url}")

        else:
            # Default to Excel
            file_path = generate_excel(rows)
            automation_context.store_file_result(file_path)
            automation_context.mark_run_success("✓ Excel Report generated and stored.")
        
    except Exception as e:
        automation_context.mark_run_failed(f"⚠ Processing failed: {e}")


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        # Running with Speckle Automate arguments
        execute_automate_function(automate_function, FunctionInputs)
    else:
        # Test authentication and Excel export with real model data
        from specklepy.transports.server import ServerTransport
        from specklepy.api import operations
        
        print("=== Testing Speckle Authentication ===")
        try:
            client = get_client()
            user = client.active_user.get()
            print(f"✓ Logged in as {user.name} on {client.url}")
            
            # Fetch model data and test Excel export
            print("\n=== Testing Excel Export with Model Data ===")
            PROJECT_ID = "08c875bbe4"
            MODEL_ID = "7631638073"
            
            model = client.model.get(MODEL_ID, PROJECT_ID)
            print(f"✓ Model: {model.name}")
            
            versions = client.version.get_versions(MODEL_ID, PROJECT_ID, limit=1)
            latest_version = versions.items[0]
            print(f"  Latest version: {latest_version.id}")
            
            transport = ServerTransport(client=client, stream_id=PROJECT_ID)
            model_data = operations.receive(latest_version.referenced_object, transport)
            
            # Test Extraction
            rows = extract_capsule_areas(model_data)
            
            if rows:
                print(f"\n✓ Found {len(rows)} rows.")
                # Test Excel Generation locally
                output = generate_excel(rows)
                print(f"  Excel generated: {output}")
            else:
                print("\n⚠ Script ran, but no mesh area data was found.")
            
        except ValueError as e:
            print(f"✗ Authentication failed: {e}")
            print("  Make sure SPECKLE_TOKEN is set in your .env file")
        except Exception as e:
            print(f"✗ Error: {e}")
            import traceback
            traceback.print_exc()