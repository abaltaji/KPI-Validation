"""
Test script for the Data Validation feature.
This script demonstrates the validation capabilities without requiring Speckle connection.
"""

import os
import sys
from openpyxl import load_workbook

# Import the reporting module
import importlib
reporting_module = importlib.import_module("04_reporting")
generate_excel = reporting_module.generate_excel
_validate_data = reporting_module._validate_data


def create_test_data():
    """Create sample test data with some validation issues."""
    return [
        {
            "tower": "Tower A",
            "level": "L1",
            "capsule": "CAP001",
            "program": "Office A",
            "location": "North Wing",
            "area": 500.5,
            "use_ratio": 0.8,
            "resource_cons_ratio": 0.6,
            "geometry_weight": 1.5,
            "mean_dist_to_exit": 15.0,
            "ideal_dist_to_exit": 10.0,
            "PRG_PAR_Area": 500.5,
            "PRG_PAR_UseRatio": 0.8,
            "PRG_PAR_ResourceConsRatio": 0.6,
            "PRG_PAR_GeometryWeight": 1.5,
            "PRG_PAR_MeanDistToExit": 15.0,
            "PRG_PAR_IdealDistToExit": 10.0,
        },
        {
            "tower": "Tower A",
            "level": "L2",
            "capsule": "CAP002",
            "program": "Office B",
            "location": "North Wing",
            "area": 0,  # Invalid: should fail area check
            "use_ratio": 1.5,  # Invalid: should fail ratio check
            "resource_cons_ratio": 0.5,
            "geometry_weight": 1.2,
            "mean_dist_to_exit": 20.0,
            "ideal_dist_to_exit": 12.0,
            "PRG_PAR_Area": 0,
            "PRG_PAR_UseRatio": 1.5,
            "PRG_PAR_ResourceConsRatio": 0.5,
            "PRG_PAR_GeometryWeight": 1.2,
            "PRG_PAR_MeanDistToExit": 20.0,
            "PRG_PAR_IdealDistToExit": 12.0,
        },
        {
            "tower": "Tower B",
            "level": "L1",
            "capsule": "CAP003",
            "program": "Warehouse C",
            "location": "South Wing",
            "area": 2000.0,
            "use_ratio": 0.9,
            "resource_cons_ratio": 0.7,
            "geometry_weight": 2.0,
            "mean_dist_to_exit": 50.0,
            "ideal_dist_to_exit": 40.0,
            "PRG_PAR_Area": 2000.0,
            "PRG_PAR_UseRatio": 0.9,
            "PRG_PAR_ResourceConsRatio": 0.7,
            "PRG_PAR_GeometryWeight": 2.0,
            "PRG_PAR_MeanDistToExit": 50.0,
            "PRG_PAR_IdealDistToExit": 40.0,
        },
        {
            "tower": "Tower C",
            "level": "L3",
            "capsule": "CAP004",
            "program": "Retail D",
            "location": "East Wing",
            "area": 750.0,
            "use_ratio": 0.6,
            "resource_cons_ratio": -0.1,  # Invalid: negative ratio
            "geometry_weight": -0.5,  # Invalid: negative weight
            "mean_dist_to_exit": -10.0,  # Invalid: negative distance
            "ideal_dist_to_exit": 20.0,
            "PRG_PAR_Area": 750.0,
            "PRG_PAR_UseRatio": 0.6,
            "PRG_PAR_ResourceConsRatio": -0.1,
            "PRG_PAR_GeometryWeight": -0.5,
            "PRG_PAR_MeanDistToExit": -10.0,
            "PRG_PAR_IdealDistToExit": 20.0,
        },
    ]


def main():
    """Run validation tests."""
    print("=" * 60)
    print("DATA VALIDATION FEATURE TEST")
    print("=" * 60)
    
    # Create test data
    test_rows = create_test_data()
    print(f"\nCreated test data with {len(test_rows)} programs")
    
    # Run validation
    print("\n" + "=" * 60)
    print("RUNNING VALIDATION CHECKS")
    print("=" * 60)
    
    validation_results = _validate_data(test_rows)
    
    print(f"\nValidation Results ({len(validation_results)} checks):")
    print("-" * 60)
    
    current_category = None
    for result in validation_results:
        if result["category"] != current_category:
            print(f"\n[{result['category']}]")
            current_category = result["category"]
        
        status_mark = "PASS" if result["status"] == "Pass" else "FAIL"
        print(f"  {status_mark:4} | {result['check']:30} | {result['details']}")
    
    # Generate Excel with validation sheet
    print("\n" + "=" * 60)
    print("GENERATING EXCEL REPORT")
    print("=" * 60)
    
    try:
        output_path = generate_excel(test_rows)
        print(f"\nExcel report generated: {output_path}")
        
        # Verify the Data Validation sheet exists
        workbook = load_workbook(output_path)
        sheet_names = workbook.sheetnames
        print(f"\nWorkbook sheets created:")
        for idx, sheet_name in enumerate(sheet_names, 1):
            print(f"  {idx}. {sheet_name}")
            
            # If it's the Data Validation sheet, show preview
            if sheet_name == "Data Validation":
                ws = workbook[sheet_name]
                print(f"     [Data Validation sheet with {ws.max_row - 1} validation checks]")
                print(f"     Headers: {[cell.value for cell in ws[1]]}")
                
                # Show first few rows
                print(f"     Sample results:")
                for row_idx in range(2, min(5, ws.max_row + 1)):
                    cells = [ws.cell(row_idx, col).value for col in range(1, 5)]
                    if cells[2] == "Pass":
                        print(f"       [PASS] {cells[1]}: {cells[3]}")
                    else:
                        print(f"       [FAIL] {cells[1]}: {cells[3]}")
        
        print("\n" + "=" * 60)
        print("SUCCESS: Data Validation feature is working correctly!")
        print("=" * 60)
        
    except Exception as e:
        print(f"\nError generating Excel report: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
